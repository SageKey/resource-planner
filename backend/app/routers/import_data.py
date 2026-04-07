"""Excel import router -- upload or read from path."""

from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from pydantic import BaseModel

from ..deps import get_connector
from engines import SQLiteConnector, ROLE_KEYS
from engines.models import PORTFOLIO_ROLE_COLUMNS, ROSTER_ROLE_MAP

router = APIRouter(prefix="/import", tags=["import"])


class ImportPathRequest(BaseModel):
    file_path: str


class ImportResult(BaseModel):
    projects_imported: int = 0
    team_imported: int = 0
    errors: list[str] = []


def _parse_excel(file_path: str, conn: SQLiteConnector) -> ImportResult:
    """Parse an Excel workbook and upsert data into the database."""
    try:
        import openpyxl
    except ImportError:
        return ImportResult(errors=["openpyxl not installed"])

    p = Path(file_path)
    if not p.exists():
        return ImportResult(errors=[f"File not found: {file_path}"])

    errors = []
    projects_count = 0
    team_count = 0

    try:
        wb = openpyxl.load_workbook(str(p), data_only=True)
    except Exception as e:
        return ImportResult(errors=[f"Cannot open workbook: {e}"])

    # --- Portfolio sheet ---
    portfolio_sheet = None
    for name in ["Portfolio", "portfolio", "Projects", "projects"]:
        if name in wb.sheetnames:
            portfolio_sheet = wb[name]
            break

    if portfolio_sheet:
        for row_idx in range(2, portfolio_sheet.max_row + 1):
            try:
                pid = portfolio_sheet.cell(row=row_idx, column=1).value
                pname = portfolio_sheet.cell(row=row_idx, column=2).value
                if not pid or not pname:
                    continue

                fields = {
                    "id": str(pid).strip(),
                    "name": str(pname).strip(),
                }

                # Read standard columns (adjust column indices as needed)
                col_map = {
                    3: "type", 4: "portfolio", 5: "sponsor", 6: "health",
                    7: "pct_complete", 8: "priority", 9: "start_date",
                    10: "end_date", 11: "actual_end", 12: "team",
                    13: "pm", 14: "ba", 15: "functional_lead",
                    16: "technical_lead", 17: "developer_lead",
                    18: "tshirt_size", 19: "est_hours",
                }
                for col, field_name in col_map.items():
                    val = portfolio_sheet.cell(row=row_idx, column=col).value
                    if val is not None:
                        if field_name == "pct_complete":
                            try:
                                fields[field_name] = float(val)
                            except (ValueError, TypeError):
                                fields[field_name] = 0.0
                        elif field_name == "est_hours":
                            try:
                                fields[field_name] = float(val)
                            except (ValueError, TypeError):
                                fields[field_name] = 0.0
                        elif field_name in ("start_date", "end_date", "actual_end"):
                            if hasattr(val, "isoformat"):
                                fields[field_name] = val.isoformat()
                            else:
                                fields[field_name] = str(val)
                        else:
                            fields[field_name] = str(val)

                # Read role allocations from mapped columns
                for col_letter, (col_idx, role_key) in PORTFOLIO_ROLE_COLUMNS.items():
                    val = portfolio_sheet.cell(row=row_idx, column=col_idx).value
                    if val is not None:
                        try:
                            alloc = float(val)
                            fields[f"alloc_{role_key}"] = alloc
                        except (ValueError, TypeError):
                            pass

                err = conn.save_project(fields, is_new=True)
                if err and "UNIQUE constraint" in str(err):
                    err = conn.save_project(fields, is_new=False)
                if err:
                    errors.append(f"Project {pid}: {err}")
                else:
                    projects_count += 1
            except Exception as e:
                errors.append(f"Row {row_idx}: {e}")

    # --- Roster sheet ---
    roster_sheet = None
    for name in ["Roster", "roster", "Team", "team", "Resources"]:
        if name in wb.sheetnames:
            roster_sheet = wb[name]
            break

    if roster_sheet:
        for row_idx in range(2, roster_sheet.max_row + 1):
            try:
                name = roster_sheet.cell(row=row_idx, column=1).value
                role = roster_sheet.cell(row=row_idx, column=2).value
                if not name or not role:
                    continue

                role_key = ROSTER_ROLE_MAP.get(str(role).strip(), str(role).lower())

                weekly = roster_sheet.cell(row=row_idx, column=3).value or 0
                reserve = roster_sheet.cell(row=row_idx, column=4).value or 0

                fields = {
                    "name": str(name).strip(),
                    "role": str(role).strip(),
                    "role_key": role_key,
                    "team": str(roster_sheet.cell(row=row_idx, column=5).value or ""),
                    "vendor": str(roster_sheet.cell(row=row_idx, column=6).value or ""),
                    "classification": str(roster_sheet.cell(row=row_idx, column=7).value or ""),
                    "rate_per_hour": float(roster_sheet.cell(row=row_idx, column=8).value or 0),
                    "weekly_hrs_available": float(weekly),
                    "support_reserve_pct": float(reserve),
                    "include_in_capacity": True,
                }

                err = conn.save_roster_member(fields)
                if err:
                    errors.append(f"Member {name}: {err}")
                else:
                    team_count += 1
            except Exception as e:
                errors.append(f"Roster row {row_idx}: {e}")

    wb.close()
    return ImportResult(
        projects_imported=projects_count,
        team_imported=team_count,
        errors=errors,
    )


@router.post("/upload", response_model=ImportResult)
async def upload_excel(
    file: UploadFile = File(...),
    conn: SQLiteConnector = Depends(get_connector),
) -> ImportResult:
    """Upload an Excel file and import projects + roster."""
    import tempfile

    suffix = Path(file.filename or "upload.xlsx").suffix
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        return _parse_excel(tmp_path, conn)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


@router.post("/from-path", response_model=ImportResult)
def import_from_path(
    payload: ImportPathRequest,
    conn: SQLiteConnector = Depends(get_connector),
) -> ImportResult:
    """Import from an Excel file on the local filesystem."""
    return _parse_excel(payload.file_path, conn)
